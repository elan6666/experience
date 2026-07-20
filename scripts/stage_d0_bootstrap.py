"""Stage audited D0 calendar, identities, index history and four-universe codes."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from a_share_research.data.normalization import normalize_security_master
from a_share_research.data.providers import QueryRequest


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _request(item: dict[str, Any]) -> QueryRequest:
    return QueryRequest(
        endpoint=str(item["endpoint"]),
        params=dict(item["params"]),
        fields=tuple(item["fields"]),
        partition_key=str(item["partition_key"]),
        min_row_count=int(item.get("min_row_count", 0)),
        reject_at_row_count=int(item.get("reject_at_row_count", 5000)),
    )


def _raw_rows(root: Path, request: QueryRequest) -> tuple[dict[str, object], ...]:
    partition = root / request.endpoint / request.partition_key / request.request_hash
    manifest = _json(partition / "manifest.json")
    if manifest.get("request_hash") != request.request_hash:
        raise RuntimeError("raw manifest request hash mismatch")
    rows = tuple(
        json.loads(line)
        for line in (partition / "rows.jsonl").read_text(encoding="utf-8").splitlines()
        if line.strip()
    )
    if len(rows) != manifest.get("row_count"):
        raise RuntimeError("raw partition row count mismatch")
    if not request.min_row_count <= len(rows) < request.reject_at_row_count:
        raise RuntimeError("raw partition violates its row-count bounds")
    return rows


def _write_jsonl(path: Path, rows: tuple[dict[str, object], ...]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(".tmp")
    with temporary.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, sort_keys=True, ensure_ascii=False) + "\n")
    temporary.replace(path)


def _tech100_codes(workbook: Path) -> tuple[str, ...]:
    book = load_workbook(workbook, read_only=True, data_only=True)
    sheet = book["Top100"]
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    code_index = header.index("代码")
    rank_index = header.index("排名")
    codes: list[str] = []
    for row in rows:
        if row[rank_index] is None or row[code_index] is None:
            continue
        rank = int(row[rank_index])
        if not 1 <= rank <= 100:
            continue
        codes.append(str(row[code_index]).split(".", maxsplit=1)[0].zfill(6))
    if len(codes) != 100 or len(set(codes)) != 100:
        raise RuntimeError("Top100 workbook must yield exactly 100 unique codes")
    return tuple(codes)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--research-root", type=Path, required=True)
    parser.add_argument("--project-root", type=Path, default=Path(__file__).resolve().parents[1])
    parser.add_argument("--index-manifest", type=Path, required=True)
    parser.add_argument("--tech100-workbook", type=Path, required=True)
    parser.add_argument("--tech32-manifest", type=Path, required=True)
    parser.add_argument("--out-root", type=Path, required=True)
    parser.add_argument("--receipt", type=Path, required=True)
    args = parser.parse_args()
    raw_root = args.research_root / "data/raw/d0_v1"
    bootstrap_config = _json(args.project_root / "configs/data/source.yaml")
    bootstrap_requests = tuple(_request(item) for item in bootstrap_config["bounded_requests"])
    calendar_request = next(r for r in bootstrap_requests if r.endpoint == "trade_cal")
    stock_requests = tuple(r for r in bootstrap_requests if r.endpoint == "stock_basic")
    calendar_rows = _raw_rows(raw_root, calendar_request)
    stock_rows = tuple(row for request in stock_requests for row in _raw_rows(raw_root, request))

    raw_by_code: dict[str, dict[str, object]] = {}
    symbol_to_code: dict[str, str] = {}
    invalid_identities: list[dict[str, object]] = []
    for row in stock_rows:
        code = str(row["ts_code"])
        if not re.fullmatch(r"\d{6}\.(SH|SZ|BJ)", code):
            invalid_identities.append(
                {
                    "ts_code": code,
                    "list_date": row.get("list_date"),
                    "delist_date": row.get("delist_date"),
                    "reason": "nonstandard permanent identity",
                }
            )
            continue
        if code in raw_by_code and raw_by_code[code] != row:
            raise RuntimeError(f"conflicting stock_basic identity: {code}")
        raw_by_code[code] = row
        symbol_to_code[str(row["symbol"]).zfill(6)] = code
    security_rows = tuple(
        normalize_security_master(raw_by_code[code]).to_dict() for code in sorted(raw_by_code)
    )

    index_document = _json(args.index_manifest)
    index_requests = tuple(_request(item) for item in index_document["bounded_requests"])
    index_weight_rows = tuple(
        row
        for request in index_requests
        if request.endpoint == "index_weight"
        for row in _raw_rows(raw_root, request)
    )
    index_daily_rows = tuple(
        row
        for request in index_requests
        if request.endpoint == "index_daily"
        for row in _raw_rows(raw_root, request)
    )
    csi300 = {str(row["con_code"]) for row in index_weight_rows if row["index_code"] == "000300.SH"}
    star50 = {str(row["con_code"]) for row in index_weight_rows if row["index_code"] == "000688.SH"}
    if len(csi300) < 300 or len(star50) < 50:
        raise RuntimeError("dynamic index history yielded an implausibly small union")

    tech32_document = _json(args.tech32_manifest)
    tech32 = tuple(str(code) for code in tech32_document["tickers"])
    if len(tech32) != 32 or len(set(tech32)) != 32:
        raise RuntimeError("tech32 manifest must contain exactly 32 unique tickers")
    tech100_bare = _tech100_codes(args.tech100_workbook)
    missing_symbols = sorted(set(tech100_bare) - set(symbol_to_code))
    if missing_symbols:
        raise RuntimeError(f"Top100 codes absent from security master: {missing_symbols}")
    tech100 = tuple(symbol_to_code[code] for code in tech100_bare)

    universe_codes = tuple(sorted(csi300 | star50 | set(tech32) | set(tech100)))
    missing_identities = sorted(set(universe_codes) - set(raw_by_code))
    if missing_identities:
        raise RuntimeError(f"universe identities absent from security master: {missing_identities}")

    _write_jsonl(args.out_root / "trade_calendar.jsonl", calendar_rows)
    _write_jsonl(args.out_root / "security_master.jsonl", security_rows)
    _write_jsonl(args.out_root / "index_weight.jsonl", index_weight_rows)
    _write_jsonl(args.out_root / "index_daily.jsonl", index_daily_rows)
    for filename, values in (
        ("csi300_union_codes.json", sorted(csi300)),
        ("star50_union_codes.json", sorted(star50)),
        ("tech32_codes.json", list(tech32)),
        ("tech100_codes.json", list(tech100)),
        ("four_universe_union_codes.json", list(universe_codes)),
    ):
        (args.out_root / filename).write_text(
            json.dumps(values, indent=2, sort_keys=True) + "\n", encoding="utf-8"
        )

    outputs = {
        path.name: _sha256(path)
        for path in sorted(args.out_root.iterdir())
        if path.is_file()
    }
    receipt = {
        "schema_version": "d0_bootstrap_staging_v1",
        "status": "PASS",
        "counts": {
            "calendar_rows": len(calendar_rows),
            "security_master": len(security_rows),
            "quarantined_invalid_identities": len(invalid_identities),
            "index_weight_rows": len(index_weight_rows),
            "index_daily_rows": len(index_daily_rows),
            "csi300_union": len(csi300),
            "star50_union": len(star50),
            "tech32": len(tech32),
            "tech100": len(tech100),
            "four_universe_union": len(universe_codes),
        },
        "sources": {
            "index_manifest_sha256": _sha256(args.index_manifest),
            "tech32_manifest_sha256": _sha256(args.tech32_manifest),
            "tech100_workbook_sha256": _sha256(args.tech100_workbook),
        },
        "quarantined_invalid_identities": invalid_identities,
        "outputs": outputs,
    }
    args.receipt.parent.mkdir(parents=True, exist_ok=True)
    args.receipt.write_text(
        json.dumps(receipt, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    print(json.dumps(receipt, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
